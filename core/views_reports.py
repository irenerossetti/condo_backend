# core/views_reports.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncMonth, TruncDate
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import csv
import io
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

from .models import (
    Fee, Payment, MaintenanceRequest, Visitor, SecurityIncident,
    Reservation, Unit, User, ActivityLog, AccessLog
)
from .permissions import IsAdmin


class AdvancedReportsView(APIView):
    """Vista principal para reportes avanzados con estadísticas"""
    permission_classes = [IsAdmin]

    def get(self, request):
        report_type = request.query_params.get('type', 'overview')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Fechas por defecto: últimos 6 meses
        if not end_date:
            end_date = timezone.now()
        else:
            end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))

        if not start_date:
            start_date = end_date - timedelta(days=180)
        else:
            start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))

        if report_type == 'financial':
            return Response(self.get_financial_report(start_date, end_date))
        elif report_type == 'security':
            return Response(self.get_security_report(start_date, end_date))
        elif report_type == 'maintenance':
            return Response(self.get_maintenance_report(start_date, end_date))
        elif report_type == 'occupancy':
            return Response(self.get_occupancy_report(start_date, end_date))
        else:
            return Response(self.get_overview_report(start_date, end_date))

    def get_financial_report(self, start_date, end_date):
        """Reporte financiero detallado"""
        # Ingresos por mes
        income_by_month = Payment.objects.filter(
            paid_at__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('paid_at')
        ).values('month').annotate(
            total=Sum('amount')
        ).order_by('month')

        # Cuotas por estado
        fees_by_status = Fee.objects.filter(
            issued_at__range=[start_date, end_date]
        ).values('status').annotate(
            count=Count('id'),
            total=Sum('amount')
        )

        # Morosidad por unidad
        delinquent_units = Fee.objects.filter(
            status='OVERDUE'
        ).values(
            'unit__code', 'unit__tower'
        ).annotate(
            total_debt=Sum('amount'),
            count=Count('id')
        ).order_by('-total_debt')[:10]

        # Tasa de cobranza
        total_issued = Fee.objects.filter(
            issued_at__range=[start_date, end_date]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        total_paid = Payment.objects.filter(
            paid_at__range=[start_date, end_date]
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        collection_rate = (total_paid / total_issued * 100) if total_issued > 0 else 0

        # Métodos de pago
        payment_methods = Payment.objects.filter(
            paid_at__range=[start_date, end_date]
        ).values('method').annotate(
            count=Count('id'),
            total=Sum('amount')
        )

        return {
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'summary': {
                'total_issued': float(total_issued),
                'total_paid': float(total_paid),
                'total_pending': float(total_issued - total_paid),
                'collection_rate': round(float(collection_rate), 2)
            },
            'income_by_month': [
                {
                    'month': item['month'].strftime('%Y-%m'),
                    'total': float(item['total'])
                }
                for item in income_by_month
            ],
            'fees_by_status': [
                {
                    'status': item['status'],
                    'count': item['count'],
                    'total': float(item['total'])
                }
                for item in fees_by_status
            ],
            'delinquent_units': [
                {
                    'unit': f"{item['unit__tower']}-{item['unit__code']}",
                    'debt': float(item['total_debt']),
                    'count': item['count']
                }
                for item in delinquent_units
            ],
            'payment_methods': [
                {
                    'method': item['method'],
                    'count': item['count'],
                    'total': float(item['total'])
                }
                for item in payment_methods
            ]
        }

    def get_security_report(self, start_date, end_date):
        """Reporte de seguridad con IA"""
        # Incidentes por tipo
        incidents_by_type = SecurityIncident.objects.filter(
            detected_at__range=[start_date, end_date]
        ).values('incident_type').annotate(
            count=Count('id')
        ).order_by('-count')

        # Incidentes por severidad (si existe el campo)
        try:
            incidents_by_severity = SecurityIncident.objects.filter(
                detected_at__range=[start_date, end_date]
            ).values('severity').annotate(
                count=Count('id')
            )
        except:
            incidents_by_severity = []

        # Visitantes por día
        visitors_by_day = Visitor.objects.filter(
            entry_time__range=[start_date, end_date]
        ).annotate(
            date=TruncDate('entry_time')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')

        # Accesos no autorizados
        unauthorized_access = AccessLog.objects.filter(
            timestamp__range=[start_date, end_date],
            was_granted=False
        ).count()

        # Detecciones de IA (contamos incidentes con confidence_score)
        try:
            ai_detections = SecurityIncident.objects.filter(
                detected_at__range=[start_date, end_date]
            ).exclude(confidence_score__isnull=True).count()
        except:
            ai_detections = 0

        # Tiempo promedio de respuesta
        avg_response_time = SecurityIncident.objects.filter(
            detected_at__range=[start_date, end_date],
            resolved_at__isnull=False
        ).annotate(
            response_time=F('resolved_at') - F('detected_at')
        ).aggregate(
            avg=Avg('response_time')
        )['avg']

        return {
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'summary': {
                'total_incidents': SecurityIncident.objects.filter(
                    detected_at__range=[start_date, end_date]
                ).count(),
                'unauthorized_access': unauthorized_access,
                'ai_detections': ai_detections,
                'avg_response_time_hours': round(
                    avg_response_time.total_seconds() / 3600, 2
                ) if avg_response_time else 0
            },
            'incidents_by_type': [
                {
                    'type': item['incident_type'],
                    'count': item['count']
                }
                for item in incidents_by_type
            ],
            'incidents_by_severity': [
                {
                    'severity': item['severity'],
                    'count': item['count']
                }
                for item in incidents_by_severity
            ],
            'visitors_by_day': [
                {
                    'date': item['date'].isoformat(),
                    'count': item['count']
                }
                for item in visitors_by_day
            ]
        }

    def get_maintenance_report(self, start_date, end_date):
        """Reporte de mantenimiento"""
        # Solicitudes por estado
        requests_by_status = MaintenanceRequest.objects.filter(
            created_at__range=[start_date, end_date]
        ).values('status').annotate(
            count=Count('id')
        )

        # Solicitudes por prioridad
        requests_by_priority = MaintenanceRequest.objects.filter(
            created_at__range=[start_date, end_date]
        ).values('priority').annotate(
            count=Count('id')
        )

        # Tiempo promedio de resolución
        avg_resolution_time = MaintenanceRequest.objects.filter(
            created_at__range=[start_date, end_date],
            completed_at__isnull=False
        ).annotate(
            resolution_time=F('completed_at') - F('created_at')
        ).aggregate(
            avg=Avg('resolution_time')
        )['avg']

        # Solicitudes por mes
        requests_by_month = MaintenanceRequest.objects.filter(
            created_at__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')

        return {
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'summary': {
                'total_requests': MaintenanceRequest.objects.filter(
                    created_at__range=[start_date, end_date]
                ).count(),
                'completed': MaintenanceRequest.objects.filter(
                    created_at__range=[start_date, end_date],
                    status='COMPLETED'
                ).count(),
                'avg_resolution_days': round(
                    avg_resolution_time.total_seconds() / 86400, 2
                ) if avg_resolution_time else 0
            },
            'requests_by_status': [
                {
                    'status': item['status'],
                    'count': item['count']
                }
                for item in requests_by_status
            ],
            'requests_by_priority': [
                {
                    'priority': item['priority'],
                    'count': item['count']
                }
                for item in requests_by_priority
            ],
            'requests_by_month': [
                {
                    'month': item['month'].strftime('%Y-%m'),
                    'count': item['count']
                }
                for item in requests_by_month
            ]
        }

    def get_occupancy_report(self, start_date, end_date):
        """Reporte de ocupación de áreas comunes"""
        # Reservas por área
        reservations_by_area = Reservation.objects.filter(
            start_time__range=[start_date, end_date]
        ).values(
            'area__name'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        # Reservas por mes
        reservations_by_month = Reservation.objects.filter(
            start_time__range=[start_date, end_date]
        ).annotate(
            month=TruncMonth('start_time')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')

        # Tasa de ocupación por área
        from .models import CommonArea
        areas_usage = []
        for area in CommonArea.objects.filter(is_active=True):
            total_reservations = Reservation.objects.filter(
                area=area,
                start_time__range=[start_date, end_date]
            ).count()
            
            areas_usage.append({
                'area': area.name,
                'reservations': total_reservations,
                'capacity': area.capacity
            })

        return {
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'summary': {
                'total_reservations': Reservation.objects.filter(
                    start_time__range=[start_date, end_date]
                ).count(),
                'active_areas': CommonArea.objects.filter(is_active=True).count()
            },
            'reservations_by_area': [
                {
                    'area': item['area__name'],
                    'count': item['count']
                }
                for item in reservations_by_area
            ],
            'reservations_by_month': [
                {
                    'month': item['month'].strftime('%Y-%m'),
                    'count': item['count']
                }
                for item in reservations_by_month
            ],
            'areas_usage': areas_usage
        }

    def get_overview_report(self, start_date, end_date):
        """Reporte general con todos los indicadores"""
        return {
            'period': {
                'start': start_date.isoformat(),
                'end': end_date.isoformat()
            },
            'financial': self.get_financial_report(start_date, end_date)['summary'],
            'security': self.get_security_report(start_date, end_date)['summary'],
            'maintenance': self.get_maintenance_report(start_date, end_date)['summary'],
            'occupancy': self.get_occupancy_report(start_date, end_date)['summary']
        }



class ExportReportView(APIView):
    """Vista para exportar reportes en diferentes formatos"""
    permission_classes = [IsAdmin]
    
    def options(self, request, *args, **kwargs):
        """Handle OPTIONS request"""
        return Response(status=200)

    def get(self, request):
        report_type = request.query_params.get('type', 'financial')
        format_type = request.query_params.get('format', 'pdf')  # pdf, excel, csv
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Obtener datos del reporte
        reports_view = AdvancedReportsView()
        reports_view.request = request
        
        if not end_date:
            end_date = timezone.now()
        else:
            end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))

        if not start_date:
            start_date = end_date - timedelta(days=180)
        else:
            start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))

        if report_type == 'financial':
            data = reports_view.get_financial_report(start_date, end_date)
        elif report_type == 'security':
            data = reports_view.get_security_report(start_date, end_date)
        elif report_type == 'maintenance':
            data = reports_view.get_maintenance_report(start_date, end_date)
        else:
            data = reports_view.get_occupancy_report(start_date, end_date)

        if format_type == 'pdf':
            return self.export_pdf(data, report_type)
        elif format_type == 'excel':
            return self.export_excel(data, report_type)
        else:
            return self.export_csv(data, report_type)

    def export_pdf(self, data, report_type):
        """Exportar reporte a PDF con UTF-8"""
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4CAF50'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        title_map = {
            'financial': 'Reporte Financiero',
            'security': 'Reporte de Seguridad',
            'maintenance': 'Reporte de Mantenimiento',
            'occupancy': 'Reporte de Ocupación'
        }
        
        elements.append(Paragraph(title_map.get(report_type, 'Reporte'), title_style))
        elements.append(Spacer(1, 12))
        
        # Período
        period_text = f"Período: {data['period']['start'][:10]} a {data['period']['end'][:10]}"
        elements.append(Paragraph(period_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Resumen
        if 'summary' in data:
            elements.append(Paragraph('Resumen', styles['Heading2']))
            summary_data = [[k.replace('_', ' ').title(), str(v)] 
                           for k, v in data['summary'].items()]
            summary_table = Table(summary_data, colWidths=[3.5*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#E8F5E9')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
        
        # Agregar datos detallados según el tipo de reporte
        if report_type == 'financial':
            if 'fees_by_status' in data and data['fees_by_status']:
                elements.append(Paragraph('Cuotas por Estado', styles['Heading2']))
                fees_data = [['Estado', 'Cantidad', 'Total']]
                for item in data['fees_by_status']:
                    fees_data.append([item['status'], str(item['count']), f"${item['total']:,.2f}"])
                fees_table = Table(fees_data, colWidths=[2*inch, 1.5*inch, 2*inch])
                fees_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(fees_table)
                elements.append(Spacer(1, 20))
            
            if 'delinquent_units' in data and data['delinquent_units']:
                elements.append(Paragraph('Top 10 Unidades Morosas', styles['Heading2']))
                delinquent_data = [['Unidad', 'Deuda', 'Cantidad']]
                for item in data['delinquent_units']:
                    delinquent_data.append([item['unit'], f"${item['debt']:,.2f}", str(item['count'])])
                delinquent_table = Table(delinquent_data, colWidths=[2*inch, 2*inch, 1.5*inch])
                delinquent_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F44336')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(delinquent_table)
        
        elif report_type == 'security':
            if 'incidents_by_type' in data and data['incidents_by_type']:
                elements.append(Paragraph('Incidentes por Tipo', styles['Heading2']))
                incidents_data = [['Tipo', 'Cantidad']]
                for item in data['incidents_by_type']:
                    incidents_data.append([item['type'], str(item['count'])])
                incidents_table = Table(incidents_data, colWidths=[3.5*inch, 2*inch])
                incidents_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(incidents_table)
        
        elif report_type == 'maintenance':
            if 'requests_by_status' in data and data['requests_by_status']:
                elements.append(Paragraph('Solicitudes por Estado', styles['Heading2']))
                requests_data = [['Estado', 'Cantidad']]
                for item in data['requests_by_status']:
                    requests_data.append([item['status'], str(item['count'])])
                requests_table = Table(requests_data, colWidths=[3.5*inch, 2*inch])
                requests_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9800')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(requests_table)
        
        elif report_type == 'occupancy':
            if 'reservations_by_area' in data and data['reservations_by_area']:
                elements.append(Paragraph('Reservas por Área', styles['Heading2']))
                reservations_data = [['Área', 'Cantidad']]
                for item in data['reservations_by_area']:
                    reservations_data.append([item['area'], str(item['count'])])
                reservations_table = Table(reservations_data, colWidths=[3.5*inch, 2*inch])
                reservations_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9C27B0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(reservations_table)
        
        # Construir el PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}.pdf"'
        return response

    def export_excel(self, data, report_type):
        """Exportar reporte a Excel con UTF-8"""
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.title()[:31]  # Excel limita a 31 caracteres
        
        # Estilos
        title_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        title_font = Font(bold=True, color='FFFFFF', size=16)
        header_fill = PatternFill(start_color='81C784', end_color='81C784', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        # Título
        ws['A1'] = f'Reporte {report_type.title()}'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Período
        ws['A2'] = 'Período:'
        ws['A2'].font = Font(bold=True)
        ws['B2'] = f"{data['period']['start'][:10]} a {data['period']['end'][:10]}"
        
        row = 4
        
        # Resumen
        if 'summary' in data:
            ws[f'A{row}'] = 'Resumen'
            ws[f'A{row}'].font = Font(bold=True, size=14, color='4CAF50')
            row += 1
            
            # Encabezados del resumen
            ws[f'A{row}'] = 'Métrica'
            ws[f'B{row}'] = 'Valor'
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = header_font
            row += 1
            
            for key, value in data['summary'].items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            row += 2
        
        # Agregar datos detallados según el tipo de reporte
        if report_type == 'financial':
            if 'fees_by_status' in data and data['fees_by_status']:
                ws[f'A{row}'] = 'Cuotas por Estado'
                ws[f'A{row}'].font = Font(bold=True, size=14, color='4CAF50')
                row += 1
                
                ws[f'A{row}'] = 'Estado'
                ws[f'B{row}'] = 'Cantidad'
                ws[f'C{row}'] = 'Total'
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].fill = header_fill
                    ws[f'{col}{row}'].font = header_font
                row += 1
                
                for item in data['fees_by_status']:
                    ws[f'A{row}'] = item['status']
                    ws[f'B{row}'] = item['count']
                    ws[f'C{row}'] = item['total']
                    row += 1
                
                row += 2
            
            if 'delinquent_units' in data and data['delinquent_units']:
                ws[f'A{row}'] = 'Unidades Morosas'
                ws[f'A{row}'].font = Font(bold=True, size=14, color='F44336')
                row += 1
                
                ws[f'A{row}'] = 'Unidad'
                ws[f'B{row}'] = 'Deuda'
                ws[f'C{row}'] = 'Cantidad'
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color='EF5350', end_color='EF5350', fill_type='solid')
                    ws[f'{col}{row}'].font = header_font
                row += 1
                
                for item in data['delinquent_units']:
                    ws[f'A{row}'] = item['unit']
                    ws[f'B{row}'] = item['debt']
                    ws[f'C{row}'] = item['count']
                    row += 1
        
        elif report_type == 'security':
            if 'incidents_by_type' in data and data['incidents_by_type']:
                ws[f'A{row}'] = 'Incidentes por Tipo'
                ws[f'A{row}'].font = Font(bold=True, size=14, color='2196F3')
                row += 1
                
                ws[f'A{row}'] = 'Tipo'
                ws[f'B{row}'] = 'Cantidad'
                for col in ['A', 'B']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color='64B5F6', end_color='64B5F6', fill_type='solid')
                    ws[f'{col}{row}'].font = header_font
                row += 1
                
                for item in data['incidents_by_type']:
                    ws[f'A{row}'] = item['type']
                    ws[f'B{row}'] = item['count']
                    row += 1
        
        elif report_type == 'maintenance':
            if 'requests_by_status' in data and data['requests_by_status']:
                ws[f'A{row}'] = 'Solicitudes por Estado'
                ws[f'A{row}'].font = Font(bold=True, size=14, color='FF9800')
                row += 1
                
                ws[f'A{row}'] = 'Estado'
                ws[f'B{row}'] = 'Cantidad'
                for col in ['A', 'B']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color='FFB74D', end_color='FFB74D', fill_type='solid')
                    ws[f'{col}{row}'].font = header_font
                row += 1
                
                for item in data['requests_by_status']:
                    ws[f'A{row}'] = item['status']
                    ws[f'B{row}'] = item['count']
                    row += 1
        
        elif report_type == 'occupancy':
            if 'reservations_by_area' in data and data['reservations_by_area']:
                ws[f'A{row}'] = 'Reservas por Área'
                ws[f'A{row}'].font = Font(bold=True, size=14, color='9C27B0')
                row += 1
                
                ws[f'A{row}'] = 'Área'
                ws[f'B{row}'] = 'Cantidad'
                for col in ['A', 'B']:
                    ws[f'{col}{row}'].fill = PatternFill(start_color='BA68C8', end_color='BA68C8', fill_type='solid')
                    ws[f'{col}{row}'].font = header_font
                row += 1
                
                for item in data['reservations_by_area']:
                    ws[f'A{row}'] = item['area']
                    ws[f'B{row}'] = item['count']
                    row += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}.xlsx"'
        return response

    def export_csv(self, data, report_type):
        """Exportar reporte a CSV con UTF-8"""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Encabezado
        writer.writerow([f'Reporte {report_type.title()}'])
        writer.writerow(['Período', f"{data['period']['start'][:10]} a {data['period']['end'][:10]}"])
        writer.writerow([])
        
        # Resumen
        if 'summary' in data:
            writer.writerow(['Resumen'])
            for key, value in data['summary'].items():
                writer.writerow([key.replace('_', ' ').title(), value])
            writer.writerow([])
        
        # Agregar datos detallados según el tipo de reporte
        if report_type == 'financial':
            if 'fees_by_status' in data:
                writer.writerow(['Cuotas por Estado'])
                writer.writerow(['Estado', 'Cantidad', 'Total'])
                for item in data['fees_by_status']:
                    writer.writerow([item['status'], item['count'], item['total']])
                writer.writerow([])
            
            if 'delinquent_units' in data:
                writer.writerow(['Unidades Morosas'])
                writer.writerow(['Unidad', 'Deuda', 'Cantidad'])
                for item in data['delinquent_units']:
                    writer.writerow([item['unit'], item['debt'], item['count']])
                writer.writerow([])
        
        elif report_type == 'security':
            if 'incidents_by_type' in data:
                writer.writerow(['Incidentes por Tipo'])
                writer.writerow(['Tipo', 'Cantidad'])
                for item in data['incidents_by_type']:
                    writer.writerow([item['type'], item['count']])
                writer.writerow([])
        
        elif report_type == 'maintenance':
            if 'requests_by_status' in data:
                writer.writerow(['Solicitudes por Estado'])
                writer.writerow(['Estado', 'Cantidad'])
                for item in data['requests_by_status']:
                    writer.writerow([item['status'], item['count']])
                writer.writerow([])
        
        elif report_type == 'occupancy':
            if 'reservations_by_area' in data:
                writer.writerow(['Reservas por Área'])
                writer.writerow(['Área', 'Cantidad'])
                for item in data['reservations_by_area']:
                    writer.writerow([item['area'], item['count']])
                writer.writerow([])
        
        # Convertir a bytes con UTF-8 BOM para Excel
        csv_content = '\ufeff' + buffer.getvalue()
        response = HttpResponse(csv_content.encode('utf-8'), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="reporte_{report_type}.csv"'
        return response
